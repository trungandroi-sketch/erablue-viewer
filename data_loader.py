"""
data_loader.py – Google Sheets connector for Erablue Viewer
Reads directly from the public Google Sheet using the gviz CSV endpoint.
No API key required. Data is cached for 5 minutes (configurable).
"""
import io
import datetime
import ssl
import urllib.parse
import urllib.request
import openpyxl

import pandas as pd
import streamlit as st

# ─── Configuration ────────────────────────────────────────────────────────────
SHEET_ID = "17FvQ8YaVVV4U158yhGM8hWFNDfI6czuL1wPCUDMteOU"
CACHE_TTL = 86400  # seconds (24 hours)

SHEET_TABS = [
    "Erablue Existing",
    "Sellout Fixture",
    "Banner",
    "2 lantai",
    "Reklame store",
    "Fixture principle",
]

# SSL context – use certifi certs if available, fallback to system defaults
try:
    import certifi
    _SSL = ssl.create_default_context(cafile=certifi.where())
except ImportError:
    _SSL = ssl.create_default_context()


# Global in-memory cache for XLSX bytes to avoid downloading it multiple times
_XLSX_BYTES = None


def _download_xlsx() -> bytes:
    global _XLSX_BYTES
    if _XLSX_BYTES is None:
        url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=30, context=_SSL) as r:
            _XLSX_BYTES = r.read()
    return _XLSX_BYTES


# ─── Public API ───────────────────────────────────────────────────────────────
@st.cache_data(ttl=CACHE_TTL, show_spinner=False)
def load_sheet(sheet_name: str) -> pd.DataFrame:
    """Load one sheet; cached for CACHE_TTL seconds. Normalizes column names to be unique."""
    xlsx_data = _download_xlsx()
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_data), data_only=True)
    
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in the Google Sheet.")
        
    sheet = wb[sheet_name]
    
    # Read all rows from sheet
    rows = []
    for r in range(1, sheet.max_row + 1):
        rows.append([sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)])
        
    # Custom parser for "Erablue Existing"
    if sheet_name == "Erablue Existing":
        r1 = rows[0] if len(rows) > 0 else []
        r2 = rows[1] if len(rows) > 1 else []
        
        headers = []
        for i in range(len(r1)):
            p = str(r1[i]).strip() if r1[i] is not None else ""
            c = str(r2[i]).strip() if r2[i] is not None else ""
            if p and c:
                if p.lower() == c.lower() or c == "":
                    headers.append(p)
                else:
                    headers.append(f"{p} {c}")
            elif p:
                headers.append(p)
            elif c:
                headers.append(c)
            else:
                headers.append(f"Col_{i}")
                
        data_rows = rows[2:] if len(rows) > 2 else []
        df = pd.DataFrame(data_rows, columns=headers)
        df = df.dropna(how="all", axis=1)
        id_col = next((c for c in df.columns if "ID" in c and "Cửa hàng" in c), None)
        if id_col:
            df = df[df[id_col].notna() & (df[id_col].astype(str).str.strip() != "")]

    # Custom parser for "Sellout Fixture"
    elif sheet_name == "Sellout Fixture":
        headers = [str(val).strip() if val is not None else f"Col_{i}" for i, val in enumerate(rows[2])]
        data_rows = rows[4:] if len(rows) > 4 else []
        df = pd.DataFrame(data_rows, columns=headers)
        df = df.dropna(how="all", axis=1)
        id_col = next((c for c in df.columns if "ID" in c or "Id" in c or "id" in c.lower()), None)
        if id_col:
            df = df[df[id_col].notna() & (df[id_col].astype(str).str.strip() != "")]

    # Custom parser for "Banner"
    elif sheet_name == "Banner":
        r3 = rows[2] if len(rows) > 2 else []
        r4 = rows[3] if len(rows) > 3 else []
        headers = []
        for i in range(len(r3)):
            p = str(r3[i]).strip() if r3[i] is not None else ""
            c = str(r4[i]).strip() if r4[i] is not None else ""
            if p and c:
                if p.lower() == c.lower() or c == "":
                    headers.append(p)
                else:
                    headers.append(f"{p} {c}")
            elif p:
                headers.append(p)
            elif c:
                headers.append(c)
            else:
                headers.append(f"Col_{i}")
        data_rows = rows[4:] if len(rows) > 4 else []
        df = pd.DataFrame(data_rows, columns=headers)
        df = df.dropna(how="all", axis=1)
        id_col = next((c for c in df.columns if "ID" in c or "Id" in c or "id" in c.lower()), None)
        if id_col:
            df = df[df[id_col].notna() & (df[id_col].astype(str).str.strip() != "")]

    # Custom parser for "2 lantai"
    elif sheet_name == "2 lantai":
        headers = [str(val).strip() if val is not None else f"Col_{i}" for i, val in enumerate(rows[0])]
        data_rows = rows[1:] if len(rows) > 1 else []
        df = pd.DataFrame(data_rows, columns=headers)
        df = df.dropna(how="all", axis=1)
        id_col = next((c for c in df.columns if "ID" in c or "Id" in c or "id" in c.lower()), None)
        if id_col:
            df = df[df[id_col].notna() & (df[id_col].astype(str).str.strip() != "")]
            
    # Custom parser for "Reklame store"
    elif sheet_name == "Reklame store":
        r3 = rows[2] if len(rows) > 2 else []
        r4 = rows[3] if len(rows) > 3 else []
        
        # Forward fill groups in r3
        r3_filled = []
        current = None
        for val in r3:
            if val is not None and str(val).strip() != "":
                current = str(val).strip()
            r3_filled.append(current)
            
        headers = []
        for i in range(len(r3)):
            g = r3_filled[i]
            b = str(r4[i]).strip() if r4[i] is not None else ""
            if g and b:
                headers.append(f"{g} - {b}")
            elif g:
                headers.append(g)
            elif b:
                headers.append(b)
            else:
                headers.append(f"Col_{i}")
                
        data_rows = rows[4:] if len(rows) > 4 else []
        df = pd.DataFrame(data_rows, columns=headers)
        # Drop completely empty columns
        df = df.dropna(how="all", axis=1)
        # Drop row if ID is null/empty
        id_col = next((c for c in df.columns if "ID" in c or "id" in c.lower()), None)
        if id_col:
            df = df[df[id_col].notna() & (df[id_col].astype(str).str.strip() != "")]
            
    # Custom parser for "Fixture principle"
    elif sheet_name == "Fixture principle":
        r2 = rows[1] if len(rows) > 1 else []
        r3 = rows[2] if len(rows) > 2 else []
        
        # Forward fill groups in r2
        r2_filled = []
        current = None
        for val in r2:
            if val is not None and str(val).strip() != "":
                current = str(val).strip()
            r2_filled.append(current)
            
        headers = []
        for i in range(len(r2)):
            g = r2_filled[i]
            b = str(r3[i]).strip() if r3[i] is not None else ""
            if g and b:
                headers.append(f"{g} - {b}")
            elif g:
                headers.append(g)
            elif b:
                headers.append(b)
            else:
                headers.append(f"Col_{i}")
                
        data_rows = rows[4:] if len(rows) > 4 else []  # Skip row 4 (totals row)
        df = pd.DataFrame(data_rows, columns=headers)
        # Drop completely empty columns
        df = df.dropna(how="all", axis=1)
        # Drop row if ID is null/empty
        id_col = next((c for c in df.columns if "ID" in c or "id" in c.lower()), None)
        if id_col:
            df = df[df[id_col].notna() & (df[id_col].astype(str).str.strip() != "")]
            
    # Standard parser for other sheets
    else:
        xlsx_file = io.BytesIO(xlsx_data)
        df = pd.read_excel(xlsx_file, sheet_name=sheet_name, header=0, dtype=str)
        df = df.dropna(how="all")
        
    # Drop columns that are completely empty / Unnamed
    df = df.loc[:, ~df.columns.astype(str).str.startswith("Unnamed")]
    df = df.loc[:, df.columns.astype(str).str.strip() != "."]
    
    # Normalize column names: collapse spaces and ensure uniqueness
    seen = {}
    new_cols = []
    for c in df.columns:
        normalized = " ".join(str(c).split())
        if normalized in seen:
            seen[normalized] += 1
            new_cols.append(f"{normalized}.{seen[normalized]}")
        else:
            seen[normalized] = 0
            new_cols.append(normalized)
    df.columns = new_cols
    return df


@st.cache_data(ttl=CACHE_TTL, show_spinner=False)
def load_erablue() -> pd.DataFrame:
    """
    Load 'Erablue Existing' and coerce numeric columns.
    Drops trailing summary rows (rows without a valid ID).
    """
    df = load_sheet("Erablue Existing")
    # Drop rows with no ID (summary rows at bottom)
    id_col = next((c for c in df.columns if "ID" in c and "Cửa hàng" in c), None)
    if id_col:
        df = df[df[id_col].notna() & (df[id_col].astype(str).str.strip() != "")]
    # Drop fully unnamed/empty columns
    df = df.loc[:, ~df.columns.str.startswith("Unnamed")]
    df = df.loc[:, df.columns.str.strip() != "."]
    # Normalize location column values casing (e.g. Jawa Barat vs Jawa barat)
    for loc_col in ["Khu vực", "Tỉnh/Thành phố", "Tỉnh/Thành phố (Rút gọn)"]:
        if loc_col in df.columns:
            df[loc_col] = df[loc_col].apply(lambda x: str(x).strip().title() if pd.notna(x) and str(x).strip() != "" else x)
            
    # Coerce numeric columns (>30% parseable as float → convert)
    for col in df.columns:
        series = df[col]
        if hasattr(series, "dtype") and pd.api.types.is_object_dtype(series):
            converted = pd.to_numeric(series, errors="coerce")
            if converted.notna().sum() > len(df) * 0.3:
                df[col] = converted
    return df


def refresh():
    """Clear all cached data – next access will re-fetch from Google Sheets."""
    global _XLSX_BYTES
    _XLSX_BYTES = None
    st.cache_data.clear()


def last_refresh_label() -> str:
    now = datetime.datetime.now()
    return now.strftime("Cập nhật lúc %H:%M ngày %d/%m/%Y")
